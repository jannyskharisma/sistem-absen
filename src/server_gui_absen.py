# server_gui_absen.py (dengan validasi agar tidak absen dua kali)
import socket
import select
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import threading
import openpyxl
import os
import time as t

HOST = '127.0.0.1'
PORT = 9000

server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
server_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
server_socket.bind((HOST, PORT))
server_socket.listen()

sockets_list = [server_socket]
absensi = {}

window = tk.Tk()
window.title("Server Absensi")
window.geometry("600x480")

autosave = tk.BooleanVar()

label = tk.Label(window, text="Daftar Absen:", font=("Arial", 12, "bold"))
label.pack(pady=5)

tree = ttk.Treeview(window, columns=("nama", "waktu"), show="headings")
tree.heading("nama", text="Nama")
tree.heading("waktu", text="Waktu Absen")
tree.column("nama", anchor=tk.CENTER)
tree.column("waktu", anchor=tk.CENTER)
tree.pack(fill=tk.BOTH, expand=True)

btn_frame = tk.Frame(window)
btn_frame.pack(pady=10)

excel_path = None

load_excel_btn = tk.Button(btn_frame, text="Load Excel", command=lambda: pilih_excel())
load_excel_btn.pack(side=tk.LEFT, padx=5)

save_excel_btn = tk.Button(btn_frame, text="Simpan ke Excel", command=lambda: simpan_excel())
save_excel_btn.pack(side=tk.LEFT, padx=5)

autosave_check = tk.Checkbutton(btn_frame, text="Auto Save", variable=autosave)
autosave_check.pack(side=tk.LEFT, padx=5)

def pilih_excel():
    global excel_path
    path = filedialog.askopenfilename(title="Pilih File Excel", filetypes=[("Excel files", "*.xlsx")])
    if path:
        excel_path = path
        messagebox.showinfo("Berhasil", f"Excel dipilih:\n{excel_path}")

def simpan_excel():
    if not excel_path:
        messagebox.showwarning("Belum Memilih File", "Silakan pilih file Excel dulu.")
        return
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        for nama, waktu in absensi.items():
            tanggal = datetime.now().strftime("%d %B %Y")
            # Cek apakah nama dan tanggal sudah pernah dicatat
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == nama and row[1] == tanggal:
                return  # Sudah ada, jangan duplikat

        ws.append([nama, tanggal, waktu])
        wb.save(excel_path)
        messagebox.showinfo("Berhasil", "Data berhasil disimpan ke Excel.")
    except:
        messagebox.showerror("Gagal", "Tidak bisa menyimpan ke file Excel.")

def update_tree():
    tree.delete(*tree.get_children())
    for nama, waktu in absensi.items():
        tree.insert("", "end", values=(nama, waktu))

def server_loop():
    while True:
        read_sockets, _, _ = select.select(sockets_list, [], [], 0.1)
        for sock in read_sockets:
            if sock == server_socket:
                client, addr = server_socket.accept()
                sockets_list.append(client)
            else:
                try:
                    data = sock.recv(1024).decode()
                    if data not in murid_terdaftar:
                        sock.send("❌ Nama tidak terdaftar di database murid.".encode())
                        t.sleep(0.2)  # beri waktu client membaca pesan
                        sockets_list.remove(sock)
                        sock.close()
                        continue
                    pesan = ""
                    tanggal = datetime.now().strftime("%d %B %Y")
                    waktu_sekarang = datetime.now().strftime("%H:%M:%S")
                    sudah_ada = False

                    if excel_path:
                        try:
                            wb = openpyxl.load_workbook(excel_path)
                            ws = wb.active
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                if row[0] == data and row[1] == tanggal:
                                    sudah_ada = True
                                    break
                        except:
                            pass

                    if sudah_ada:
                        pesan = "⚠️ Kamu sudah absen sebelumnya."
                    else:
                        absensi[data] = waktu_sekarang
                        try:
                            wb = openpyxl.load_workbook(excel_path)
                            ws = wb.active
                            ws.append([data, tanggal, waktu_sekarang])
                            wb.save(excel_path)
                        except:
                            pass
                        pesan = "✅ Absen dicatat. Terima kasih!"
                        update_tree()

                    sock.send(pesan.encode())
                    pesan = ""
                    if data:
                        if data in absensi:
                            pesan = "⚠️ Kamu sudah absen sebelumnya."

                    # Cek Excel sebelum lanjut
                    if pesan == "" and excel_path:
                        try:
                            wb = openpyxl.load_workbook(excel_path)
                            ws = wb.active
                            tanggal = datetime.now().strftime("%d %B %Y")
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                if row[0] == data and row[1] == tanggal:
                                    pesan = "⚠️ Kamu sudah absen sebelumnya."
                                    break
                        except:
                            pass
                            sock.send("⚠️ Kamu sudah absen sebelumnya.".encode())
                            continue

                        if pesan == "":
                            now = datetime.now().strftime("%H:%M:%S")
                            absensi[data] = now
                            pesan = "✅ Absen dicatat. Terima kasih!"

                        sock.send(pesan.encode())

                        if autosave.get():
                            simpan_excel()

                        with open("data/daftar_absen.txt", "w") as f:
                            for nama, waktu in absensi.items():
                                f.write(f"{nama} - {waktu}\n")

                        update_tree()
                    else:
                        sockets_list.remove(sock)
                        sock.close()
                except:
                    sockets_list.remove(sock)
                    sock.close()

def auto_reset():
    while True:
        now = datetime.now()
        if now.strftime("%H:%M") == "00:00":
            absensi.clear()
            update_tree()
            with open("data/daftar_absen.txt", "w") as f:
                f.write("")
            t.sleep(60)
        t.sleep(1)

def load_murid_db(path="murid.xlsx"):
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        return set(row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0])
    except:
        return set()

murid_terdaftar = load_murid_db()

threading.Thread(target=server_loop, daemon=True).start()
threading.Thread(target=auto_reset, daemon=True).start()

window.mainloop()